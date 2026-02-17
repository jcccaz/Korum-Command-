"""
KorumOS File Processor
Handles uploaded files: images (base64), PDFs, DOCX, XLSX → extracted text.
"""
import base64
import io
import os

ALLOWED_IMAGE_TYPES = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
ALLOWED_DOC_TYPES = {'.pdf', '.docx', '.xlsx'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def process_uploaded_file(file_storage):
    """
    Takes a werkzeug FileStorage object.
    Returns dict with type, filename, base64/mime_type (images), or extracted_text (documents).
    """
    filename = file_storage.filename or "unknown"
    ext = os.path.splitext(filename)[1].lower()
    content = file_storage.read()

    if len(content) > MAX_FILE_SIZE:
        raise ValueError(f"File {filename} exceeds 10MB limit")

    if ext in ALLOWED_IMAGE_TYPES:
        return _process_image(content, filename, ext)
    elif ext == '.pdf':
        return _process_pdf(content, filename)
    elif ext == '.docx':
        return _process_docx(content, filename)
    elif ext == '.xlsx':
        return _process_xlsx(content, filename)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _process_image(content, filename, ext):
    mime_map = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif', '.webp': 'image/webp'
    }
    b64 = base64.standard_b64encode(content).decode('utf-8')
    return {
        'type': 'image',
        'filename': filename,
        'base64': b64,
        'mime_type': mime_map.get(ext, 'image/jpeg'),
        'extracted_text': None
    }


def _process_pdf(content, filename):
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(content))
    text_parts = []
    for page in reader.pages[:50]:  # Cap at 50 pages
        text_parts.append(page.extract_text() or '')
    return {
        'type': 'document',
        'filename': filename,
        'base64': None,
        'mime_type': None,
        'extracted_text': '\n'.join(text_parts)[:20000]
    }


def _process_docx(content, filename):
    from docx import Document
    doc = Document(io.BytesIO(content))
    text = '\n'.join(p.text for p in doc.paragraphs)
    return {
        'type': 'document',
        'filename': filename,
        'base64': None,
        'mime_type': None,
        'extracted_text': text[:20000]
    }


def _process_xlsx(content, filename):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames[:5]:  # Cap at 5 sheets
        ws = wb[sheet_name]
        text_parts.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(max_row=200, values_only=True):
            text_parts.append(' | '.join(str(c) if c is not None else '' for c in row))
    return {
        'type': 'document',
        'filename': filename,
        'base64': None,
        'mime_type': None,
        'extracted_text': '\n'.join(text_parts)[:20000]
    }
